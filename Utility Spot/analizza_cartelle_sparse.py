#!/usr/bin/env python3
"""
Lista rapida delle cartelle Desktop/Downloads sparse senza analisi pesanti.
"""

# --- IRC shared bootstrap ---
# Rende disponibili i moduli in Python/shared/ senza dipendere da PYTHONPATH.
# Saltato se eseguito da bundle PyInstaller (sys.frozen=True): in quel caso
# i moduli sono gia' inclusi nel bundle.
import sys as _sys
from pathlib import Path as _Path
if not getattr(_sys, 'frozen', False):
    _shared = _Path.home() / "Library/CloudStorage/Dropbox/Documenti_IRC/Python/shared"
    if str(_shared) not in _sys.path:
        _sys.path.insert(0, str(_shared))
# --- end IRC shared bootstrap ---

APP_NAME = "AnalizzaCartelleSparse"
VERSION  = "1.1.0"

from irc_paths import app_output_dir
from irc_logging import setup_app_logger

import os
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

log = setup_app_logger(APP_NAME, also_to_console=True)

def categorize_location(path):
    """Categorizza la posizione della cartella."""
    path_str = str(path)
    
    if 'Dropbox' in path_str:
        if 'Il mio Mac' in path_str or 'My Mac' in path_str:
            return 'Dropbox - Backup Mac'
        elif 'saved' in path_str:
            return 'Dropbox - Saved'
        else:
            return 'Dropbox'
    elif 'GoogleDrive' in path_str:
        return 'Google Drive'
    elif 'iCloud' in path_str or 'CloudDocs' in path_str:
        return 'iCloud'
    elif '/Volumes/' in path_str:
        return 'Disco Esterno'
    elif 'Application Support' in path_str:
        return 'App Support'
    elif path_str.startswith('/Users/') and path_str.count('/') == 3:
        return '🏠 HOME UTENTE (PRINCIPALE)'
    else:
        return 'Altro'

def find_folders_fast():
    """Trova cartelle senza timeout."""
    print("\n🔍 Ricerca cartelle...\n")
    
    folders = []
    folders_to_find = ['Desktop', 'Scrivania', 'Downloads', 'Scaricati']
    
    # 1. Home directory utenti (escludi silvia)
    print("  📁 Home directory utenti...")
    users_dir = Path('/Users')
    for user_dir in users_dir.iterdir():
        if user_dir.is_dir() and user_dir.name not in ['Shared', '.localized', 'silvia']:
            for folder_name in folders_to_find:
                folder_path = user_dir / folder_name
                if folder_path.exists():
                    folders.append(str(folder_path))
                    print(f"    ✓ {folder_path}")
    
    # 2. CloudStorage (limitato a 1 livello, escludi silvia)
    print("\n  ☁️  Cloud Storage...")
    for user_dir in users_dir.iterdir():
        if user_dir.is_dir() and user_dir.name not in ['Shared', '.localized', 'silvia']:
            cloud_dir = user_dir / 'Library' / 'CloudStorage'
            if cloud_dir.exists():
                try:
                    # Solo primo livello sotto ogni servizio cloud
                    for cloud_service in cloud_dir.iterdir():
                        if cloud_service.is_dir():
                            for folder_name in folders_to_find:
                                # Cerca direttamente
                                direct = cloud_service / folder_name
                                if direct.exists() and direct.is_dir():
                                    folders.append(str(direct))
                                    print(f"    ✓ {direct}")
                                
                                # Cerca in sottocartelle comuni
                                for subdir in ['saved', 'Data']:
                                    sub_path = cloud_service / subdir / folder_name
                                    if sub_path.exists() and sub_path.is_dir():
                                        folders.append(str(sub_path))
                                        print(f"    ✓ {sub_path}")
                                
                                # Cerca backup Mac
                                try:
                                    for item in cloud_service.iterdir():
                                        if item.is_dir() and ('Mac' in item.name or 'iMac' in item.name):
                                            mac_folder = item / folder_name
                                            if mac_folder.exists() and mac_folder.is_dir():
                                                folders.append(str(mac_folder))
                                                print(f"    ✓ {mac_folder}")
                                except:
                                    pass
                except Exception as e:
                    print(f"    ⚠️  Errore: {e}")
    
    print(f"\n✅ Trovate {len(folders)} cartelle\n")
    return folders

def create_simple_report(folders, output_path):
    """Crea report Excel semplice con rilevamento symlink."""
    print("📝 Creazione report...\n")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cartelle Trovate"
    
    # Header
    headers = ['#', 'Nome', 'Percorso Completo', 'Tipo', 'Link → Destinazione', 'Categoria', 'Azione Suggerita']
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Analizza e ordina
    analyzed = []
    for folder in folders:
        folder_path = Path(folder)
        category = categorize_location(folder)
        name = os.path.basename(folder)
        
        # Verifica se è symlink
        is_symlink = folder_path.is_symlink()
        link_target = ""
        file_type = "📁 Cartella"
        
        if is_symlink:
            try:
                link_target = str(folder_path.readlink())
                # Risolvi path relativo
                if not link_target.startswith('/'):
                    link_target = str((folder_path.parent / link_target).resolve())
                file_type = "🔗 Symlink"
            except:
                link_target = "Errore lettura"
        
        # Azione suggerita
        if is_symlink:
            action = '🗑️  CANCELLA - Symlink inutile (non perdi dati)'
        elif category == '🏠 HOME UTENTE (PRINCIPALE)':
            action = '✅ MANTIENI - Cartella principale'
        elif 'Backup Mac' in category:
            action = '⚠️  VERIFICA - Backup vecchio Mac, probabilmente da cancellare'
        elif category in ['Dropbox - Saved', 'Dropbox']:
            action = '⚠️  VERIFICA - Backup/duplicato, controlla contenuto'
        elif category == 'App Support':
            action = '❌ IGNORA - File di sistema/app'
        else:
            action = '❓ VERIFICA - Controlla contenuto'
        
        analyzed.append({
            'name': name,
            'path': folder,
            'type': file_type,
            'link_target': link_target,
            'category': category,
            'action': action,
            'is_symlink': is_symlink
        })
    
    # Ordina: symlink prima, poi per categoria
    analyzed.sort(key=lambda x: (not x['is_symlink'], x['category'], x['path']))
    
    # Scrivi dati
    for idx, item in enumerate(analyzed, 2):
        ws.cell(row=idx, column=1, value=idx-1)
        ws.cell(row=idx, column=2, value=item['name'])
        ws.cell(row=idx, column=3, value=item['path'])
        ws.cell(row=idx, column=4, value=item['type'])
        ws.cell(row=idx, column=5, value=item['link_target'])
        ws.cell(row=idx, column=6, value=item['category'])
        ws.cell(row=idx, column=7, value=item['action'])
        
        # Colori
        if item['is_symlink']:
            fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")  # Viola chiaro
        elif item['category'] == '🏠 HOME UTENTE (PRINCIPALE)':
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif 'Backup Mac' in item['category']:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif item['category'] in ['Dropbox - Saved', 'Dropbox']:
            fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            fill = None
        
        if fill:
            for col in range(1, 8):
                ws.cell(row=idx, column=col).fill = fill
    
    # Auto-width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    print(f"✅ Report salvato: {output_path}\n")

def main():
    print("\n" + "="*70)
    print("  LISTA RAPIDA CARTELLE DESKTOP/DOWNLOADS")
    print("="*70)
    log.info(f"Avviato v{VERSION}")

    folders = find_folders_fast()
    
    if not folders:
        print("❌ Nessuna cartella trovata!")
        return
    
    output_path = str(app_output_dir(APP_NAME) / f"Cartelle_Lista_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    create_simple_report(folders, output_path)
    log.info(f"Report salvato: {output_path}  ({len(folders)} cartelle)")

    print("="*70)
    print(f"📊 Trovate {len(folders)} cartelle")
    print(f"💾 Report Excel in Documents/output/AnalizzaCartelleSparse/")
    print("="*70 + "\n")
    import subprocess
    subprocess.run(["open", output_path])
    
    print("🔍 LEGENDA COLORI:")
    print("  🟣 Viola = SYMLINK da cancellare (non perdi dati)")
    print("  🟢 Verde = Cartella principale da MANTENERE")
    print("  🔴 Rosso = Backup vecchi Mac da VERIFICARE/CANCELLARE")
    print("  🟡 Giallo = Dropbox duplicati da VERIFICARE")
    print()

if __name__ == "__main__":
    main()