#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Inventario_Hardware_iMac_Completo_1.2
Autore: Ignazio Rusconi-Clerici
Data: 2025-10-23

Descrizione:
Inventario completo di hardware, dischi, rete e software su macOS.
Versione corretta che ripristina le fonti dati originali e il formato
Excel coerente con l’analisi Cloud (intestazioni blu, autosize colonne).
"""

import os
import sys
import argparse
import subprocess
import plistlib
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- Percorsi ---
BASE_DOWNLOAD_DIR = os.path.expanduser("~/Documents/download")
BASE_LOG_DIR      = os.path.expanduser("~/Documents/log")

def make_output_paths(machine_name: str):
    """Costruisce i percorsi di output usando il nome macchina."""
    safe = "".join(c if c.isalnum() or c in " _-" else "_" for c in machine_name).strip()
    safe = safe or "iMac"
    base = f"{safe}_{datetime.now():%Y-%m-%d_%H-%M-%S}"
    os.makedirs(BASE_DOWNLOAD_DIR, exist_ok=True)
    os.makedirs(BASE_LOG_DIR, exist_ok=True)
    return (os.path.join(BASE_DOWNLOAD_DIR, f"{base}.xlsx"),
            os.path.join(BASE_LOG_DIR,      f"{base}.log"))

# --- Utility ---
def run_cmd(cmd):
    try:
        result = subprocess.run(cmd, shell=True, text=True, stdout=subprocess.PIPE, stderr=subprocess.DEVNULL)
        return result.stdout.strip()
    except Exception:
        return "N/D"

# --- Sezione Hardware ---
def collect_hardware():
    hw = []
    def add(voce, valore, descrizione):
        hw.append({"Voce": voce, "Valore": valore, "Descrizione": descrizione})

    add("Modello (hw.model)", run_cmd("sysctl -n hw.model"), "Identificatore hardware del Mac")
    add("CPU/Chip", run_cmd("sysctl -n machdep.cpu.brand_string"), "Nome del processore o SoC")
    add("Core totali", run_cmd("sysctl -n hw.physicalcpu") + " fisici / " + run_cmd("sysctl -n hw.logicalcpu") + " logici", "Numero di core fisici e logici")
    add("RAM totale (GB)", str(round(int(run_cmd("sysctl -n hw.memsize")) / (1024**3), 2)), "Memoria installata")
    add("Numero di serie", run_cmd("system_profiler SPHardwareDataType | awk '/Serial/ {print $4}'"), "Numero di serie del dispositivo")
    add("macOS versione", run_cmd("sw_vers -productVersion") + " (" + run_cmd("sw_vers -buildVersion") + ")", "Versione e build di macOS")
    add("Uptime", run_cmd("uptime | awk -F'up ' '{print $2}' | cut -d',' -f1-3"), "Tempo di attività dall’ultimo avvio")
    add("Utente attuale", run_cmd("whoami"), "Account loggato attualmente")

    gpu = run_cmd("system_profiler SPDisplaysDataType | awk -F': ' '/Chipset Model/ {print $2}' | head -1")
    if gpu:
        add("GPU", gpu, "Chip grafico principale")

    return pd.DataFrame(hw)

def collect_disks():
    """
    Elenca tutti i dischi fisici e le partizioni (diskXsY),
    incluso EFI, Recovery, APFS container, ecc.
    """
    def parse_gb(value):
        if not value or value == "—":
            return "—"
        if "GB" in value:
            return value.split("GB")[0].strip()
        if "GiB" in value:
            return value.split("GiB")[0].strip()
        return "—"

    disks = []
    list_output = run_cmd("diskutil list | grep -o 'disk[0-9s]*' | sort -u")

    for dev in list_output.splitlines():
        info = run_cmd(f"diskutil info {dev}")
        if not info:
            continue

        def find(key):
            for l in info.splitlines():
                if key in l:
                    return l.split(':', 1)[1].strip()
            return "—"

        size_raw = find("Total Size")
        free_raw = find("Free Space")

        disks.append({
            "Device": dev,
            "Interno": "Sì" if "Yes" in find("Internal") else "No",
            "SSD": "Sì" if "Yes" in find("Solid State") else "No",
            "Bus": find("Bus Protocol") or find("Device Location") or "—",
            "Volume": find("Volume Name"),
            "File system": find("File System Personality"),
            "Capacità (GB)": parse_gb(size_raw),
            "Libero (GB)": parse_gb(free_raw),
            "Mount": find("Mount Point")
        })

    return pd.DataFrame(disks)

# --- Sezione Rete ---
def collect_network():
    hw_ports = run_cmd("networksetup -listallhardwareports")
    blocks = hw_ports.strip().split("\n\n")
    rete = []
    for block in blocks:
        lines = block.splitlines()
        port, device, mac = "—","—","—"
        for l in lines:
            if l.startswith("Hardware Port:"):
                port = l.split(": ")[1].strip()
            elif l.startswith("Device:"):
                device = l.split(": ")[1].strip()
            elif l.startswith("Ethernet Address:"):
                mac = l.split(": ")[1].strip()
        if device == "—":
            continue
        ipv4 = run_cmd(f"ipconfig getifaddr {device}")
        attiva = "Sì" if ipv4 else "No"
        rete.append({
            "Interfaccia": port,
            "Attiva": attiva,
            "IPv4": ipv4 if ipv4 else "—",
            "Indirizzi": ipv4 if ipv4 else "—",
            "MAC": mac,
            "Dettagli": device
        })
    return pd.DataFrame(rete)

# --- Sezione Software ---
def collect_software():
    try:
        plist_data = plistlib.loads(subprocess.check_output("system_profiler SPApplicationsDataType -xml", shell=True))
        items = plist_data[0]["_items"]
    except Exception:
        return pd.DataFrame(columns=["Applicazione", "Versione", "Percorso"])

    software = []
    for app in items:
        software.append({
            "Applicazione": app.get("_name", "—"),
            "Versione": app.get("version", "—"),
            "Percorso": app.get("path", "—")
        })
    return pd.DataFrame(software)

# --- Formattazione Excel ---
def format_excel(path):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(border_style="thin", color="D9D9D9")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)
    wb.save(path)

# --- Main ---
def main():
    parser = argparse.ArgumentParser(description="Inventario hardware/software Mac")
    parser.add_argument(
        "--etichetta",
        default="iMac",
        help="Etichetta per i nomi di output (es. iMac BdS, iMac Gignese)."
    )
    args = parser.parse_args()

    output_xlsx, output_log = make_output_paths(args.etichetta)

    print(f"🔍 Raccolta informazioni di sistema ({args.etichetta})...")
    dataframes = {
        "Hardware": collect_hardware(),
        "Dischi": collect_disks(),
        "Rete": collect_network(),
        "Software": collect_software(),
    }

    print("💾 Generazione file Excel...")
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        for name, df in dataframes.items():
            df.to_excel(writer, index=False, sheet_name=name)
    format_excel(output_xlsx)

    # Scrivi log testuale
    log_lines = [
        f"Inventario_Hardware_iMac_Completo_1.2",
        f"Macchina: {args.etichetta}",
        f"Data esecuzione: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Excel output: {output_xlsx}",
        f"Log output: {output_log}",
    ]
    with open(output_log, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines) + "\n")

    print(f"✅ Inventario completato: {output_xlsx}")
    print(f"📝 Log: {output_log}")
    import subprocess
    subprocess.run(["open", output_xlsx])

if __name__ == "__main__":
    main()