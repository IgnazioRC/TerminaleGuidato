#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Dropbox_FileStatus_OnlyOnline_1.4
Autore: Ignazio Rusconi-Clerici + Claude
Crea un report Excel solo con i file Dropbox “Solo online”
"""

import os
import subprocess
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === CONFIGURAZIONE ===
DROPBOX_PATH = Path.home() / "Library/CloudStorage/Dropbox"
OUTPUT_FILE = Path.home() / "Desktop/Dropbox_SoloOnline.xlsx"

def is_only_online(path: Path) -> bool:
    """Restituisce True se il file è 'solo online' in Dropbox.
    Usa /usr/bin/xattr (sempre presente su macOS) senza dipendenze esterne.
    """
    try:
        result = subprocess.run(
            ["/usr/bin/xattr", "-p", "com.dropbox.attributes", str(path)],
            capture_output=True, text=True
        )
        return result.returncode == 0 and "only_on_cloud" in result.stdout
    except Exception:
        return False

def scan_dropbox(base_path: Path):
    """Scansiona solo i file 'solo online', con avanzamento ogni 100 file."""
    results = []
    totale = 0
    scansionati = 0
    STEP = 100

    # Conta veloce per avere un totale orientativo
    print("⏳ Conteggio file in corso...", flush=True)
    for root, dirs, files in os.walk(base_path):
        totale += len(files)
    print(f"📁 File totali da esaminare: {totale}", flush=True)
    print(f"   (avanzamento ogni {STEP} file)", flush=True)

    for root, dirs, files in os.walk(base_path):
        for name in files:
            full_path = Path(root) / name
            scansionati += 1
            if scansionati % STEP == 0:
                perc = scansionati * 100 // totale if totale else 0
                print(f"   [{perc:3d}%] {scansionati}/{totale} file esaminati"
                      f" — trovati: {len(results)}", flush=True)
            try:
                if is_only_online(full_path):
                    mtime = datetime.fromtimestamp(full_path.stat().st_mtime)
                    results.append([
                        str(full_path),
                        round(full_path.stat().st_size / (1024 * 1024), 2),
                        mtime
                    ])
            except Exception:
                continue

    print(f"   [100%] {scansionati}/{totale} file esaminati"
          f" — trovati: {len(results)}", flush=True)
    return results

def save_excel(data, output_path: Path):
    """Salva il report Excel"""
    df = pd.DataFrame(data, columns=["Percorso", "Dimensione (MB)", "Ultima modifica"])
    df.sort_values("Percorso", inplace=True)
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    gray = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).fill = gray
    wb.save(output_path)

def main():
    print(f"📦 Scansione Dropbox: {DROPBOX_PATH}")
    data = scan_dropbox(DROPBOX_PATH)
    print(f"📊 File 'solo online' trovati: {len(data)}")
    if not data:
        print("ℹ️  Nessun file solo online trovato. Report non generato.")
        return
    save_excel(data, OUTPUT_FILE)
    print(f"✅ Report completato: {OUTPUT_FILE}")
    subprocess.run(["open", str(OUTPUT_FILE)])

if __name__ == "__main__":
    main()