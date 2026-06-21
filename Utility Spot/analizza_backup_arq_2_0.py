#!/usr/bin/env python3
"""
Arq Backup Analyzer v2.1.0 (CLI Version - Ottimizzato)
Analizza i backup Arq 7 non criptati e genera report dettagliati.

Ottimizzazioni v2.1:
- Calcolo dimensioni opzionale (--calc-size)
- Progress feedback durante scansione
- Limite record per evitare blocchi
- Timeout per operazioni su cloud storage

Uso:
    python arq_analyzer_v2.py [percorso_backup_arq]
    python arq_analyzer_v2.py --calc-size    # Include calcolo dimensioni (lento!)
    python arq_analyzer_v2.py --max-records 50  # Limita record per cartella

Autore: Ignazio
Data: 2026-01-06
"""

# --- IRC shared bootstrap ---
# Rende disponibili i moduli in Python/shared/ senza dipendere da PYTHONPATH.
# Saltato se eseguito da bundle PyInstaller (sys.frozen=True): in quel caso
# i moduli sono gia' inclusi nel bundle.
import sys as _sys
from pathlib import Path as _Path
if not getattr(_sys, 'frozen', False):
    _shared = _Path.home() / "Library/CloudStorage/Dropbox/Documenti_IRC/Python/shared"
    if str(_shared) not in _sys.path:
        _sys.path.insert(0, str(_shared))
# --- end IRC shared bootstrap ---

APP_NAME = "AnalizzaBackupArq"
VERSION  = "2.2.0"

from irc_paths import app_output_dir
from irc_logging import setup_app_logger

import os
import sys
import json
import struct
import argparse
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import Dict, List, Optional, Any, Tuple

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
log = setup_app_logger(APP_NAME, also_to_console=True)

# Verifica dipendenze
try:
    import lz4.block
    HAS_LZ4 = True
except ImportError:
    HAS_LZ4 = False
    log.warning("Modulo lz4 non trovato. Installare con: pip install lz4")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    log.warning("Modulo openpyxl non trovato. Installare con: pip install openpyxl")

# Mappatura UUID → Nome Mac
MAC_MAPPING = {
    '7F0CEB6C-7C7C-4F90-ACC9-E2734B461CFD': 'iMac Gignese',
    'A9031814-0B56-41CD-A3F5-D61A50C7F6F7': 'iMac BdS',
    'A76961B4-2FF5-4BC5-B0A5-AF50C885C953': 'MacBook Pro'
}

# Percorso default backup Arq
DEFAULT_ARQ_PATH = Path.home() / "Library/CloudStorage/GoogleDrive-ignazio.rusconiclerici@gmail.com/Il mio Drive/Arq Backup Data"


def human_size(size_bytes: int) -> str:
    """Converte bytes in formato leggibile."""
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if abs(size_bytes) < 1024.0:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024.0
    return f"{size_bytes:.1f} PB"


def extract_user_from_path(path: str) -> str:
    """Estrae username dal percorso."""
    if '/Users/' in path:
        parts = path.split('/Users/')
        if len(parts) > 1:
            return parts[1].split('/')[0]
    return 'System'


def print_progress(message: str, end: str = '\n'):
    """Stampa messaggio di progresso."""
    print(f"     {message}", end=end, flush=True)


class ArqBackupSet:
    """Rappresenta un backup set di un singolo Mac."""
    
    def __init__(self, base_path: Path, uuid: str, name: str):
        self.base_path = base_path
        self.uuid = uuid
        self.name = name
        self.path = base_path / uuid
        
        # Dati caricati
        self.config: Dict = {}
        self.plan: Dict = {}
        self.folders: List[Dict] = []
        self.stats: Dict = {}
        
        # Opzioni
        self.max_records_per_folder = 20
        self.calculate_sizes = False
        
    def exists(self) -> bool:
        return self.path.exists()
    
    def load_config(self) -> bool:
        """Carica backupconfig.json"""
        config_path = self.path / 'backupconfig.json'
        if config_path.exists():
            try:
                with open(config_path, 'r') as f:
                    content = f.read()
                    # Rimuovi commenti /* */ che Arq usa nel JSON
                    content = re.sub(r'/\*.*?\*/', '', content, flags=re.DOTALL)
                    self.config = json.loads(content)
                    return True
            except Exception as e:
                log.error(f"Errore lettura config {self.name}: {e}")
        return False
    
    def load_plan(self) -> bool:
        """Carica backupplan.json"""
        plan_path = self.path / 'backupplan.json'
        if plan_path.exists():
            try:
                with open(plan_path, 'r') as f:
                    self.plan = json.load(f)
                    return True
            except Exception as e:
                log.error(f"Errore lettura plan {self.name}: {e}")
        return False
    
    def load_folders(self, verbose: bool = True) -> List[Dict]:
        """Carica info su tutte le cartelle backuppate."""
        self.folders = []
        backupfolders_path = self.path / 'backupfolders'
        
        if not backupfolders_path.exists():
            return self.folders
        
        if verbose:
            print_progress("Lettura cartelle...", end=' ')
        
        try:
            folder_dirs = [d for d in backupfolders_path.iterdir() if d.is_dir()]
        except Exception as e:
            log.error(f"Errore lettura directory {backupfolders_path}: {e}")
            return self.folders
        
        for folder_dir in folder_dirs:
            folder_json = folder_dir / 'backupfolder.json'
            if folder_json.exists():
                try:
                    with open(folder_json, 'r') as f:
                        folder_data = json.load(f)
                        folder_data['_uuid'] = folder_dir.name
                        folder_data['_records_path'] = folder_dir / 'backuprecords'
                        self.folders.append(folder_data)
                except Exception as e:
                    log.warning(f"Errore lettura folder {folder_dir.name}: {e}")
        
        if verbose:
            print(f"{len(self.folders)} trovate")
        
        return self.folders
    
    def get_backup_dates_fast(self, verbose: bool = True) -> List[Dict]:
        """
        Recupera le date dei backup in modo veloce.
        Legge solo i nomi dei file senza aprirli.
        """
        all_dates = []
        
        if verbose:
            print_progress("Lettura date backup...", end=' ')
        
        folder_count = 0
        for folder in self.folders:
            folder_uuid = folder.get('uuid') or folder.get('_uuid')
            folder_name = folder.get('name', 'Unknown')
            folder_path = folder.get('localPath', '')
            records_path = folder.get('_records_path')
            
            if not records_path or not records_path.exists():
                continue
            
            folder_count += 1
            
            # Trova le subdirectory (sono i prefissi del timestamp)
            try:
                subdirs = [d for d in records_path.iterdir() if d.is_dir()]
            except Exception as e:
                log.debug(f"Errore lettura {records_path}: {e}")
                continue
            
            record_count = 0
            for subdir in subdirs:
                if record_count >= self.max_records_per_folder:
                    break
                    
                try:
                    # Trova i .backuprecord in questa subdir
                    for record_file in subdir.glob('*.backuprecord'):
                        if record_count >= self.max_records_per_folder:
                            break
                        
                        try:
                            # Timestamp = nome_dir + nome_file (senza estensione)
                            dir_part = subdir.name
                            file_part = record_file.stem
                            epoch = int(dir_part + file_part)
                            dt = datetime.fromtimestamp(epoch)
                            
                            all_dates.append({
                                'folder_name': folder_name,
                                'folder_path': folder_path,
                                'folder_uuid': folder_uuid,
                                'timestamp': epoch,
                                'datetime': dt,
                                'file_path': str(record_file),
                                'file_size': 0  # Non leggiamo la dimensione per velocità
                            })
                            record_count += 1
                        except ValueError:
                            pass
                except Exception as e:
                    log.debug(f"Errore lettura subdir {subdir}: {e}")
        
        if verbose:
            print(f"{len(all_dates)} record da {folder_count} cartelle")
        
        return sorted(all_dates, key=lambda x: x['timestamp'], reverse=True)
    
    def calculate_total_size(self, verbose: bool = True) -> int:
        """
        Calcola dimensione totale del backup set.
        ATTENZIONE: Può essere molto lento su cloud storage!
        """
        if not self.path.exists():
            return 0
        
        if verbose:
            print_progress("Calcolo dimensioni (può richiedere tempo)...", end=' ')
        
        total_size = 0
        file_count = 0
        
        try:
            for f in self.path.rglob('*'):
                if f.is_file():
                    try:
                        total_size += f.stat().st_size
                        file_count += 1
                        if verbose and file_count % 1000 == 0:
                            print(f"\r     Calcolo dimensioni... {file_count} file, {human_size(total_size)}", end='', flush=True)
                    except:
                        pass
        except Exception as e:
            log.warning(f"Errore calcolo dimensioni {self.name}: {e}")
        
        if verbose:
            print(f"\r     Dimensione totale: {human_size(total_size)} ({file_count} file)    ")
        
        return total_size
    
    def calculate_stats(self, verbose: bool = True) -> Dict:
        """Calcola statistiche sul backup set."""
        self.stats = {
            'is_encrypted': self.config.get('isEncrypted', False),
            'computer_name': self.config.get('computerName', 'Unknown'),
            'backup_name': self.config.get('backupName', 'Unknown'),
            'blob_type': 'SHA256' if self.config.get('blobIdentifierType') == 2 else 'SHA1',
            'chunker_version': self.config.get('chunkerVersion', 'Unknown'),
            'folders_count': len(self.folders),
            'total_records': 0,
            'oldest_backup': None,
            'newest_backup': None,
            'total_size_bytes': 0,
            'total_size_human': 'N/D (usa --calc-size)'
        }
        
        # Calcola dimensioni solo se richiesto
        if self.calculate_sizes:
            total_size = self.calculate_total_size(verbose)
            self.stats['total_size_bytes'] = total_size
            self.stats['total_size_human'] = human_size(total_size)
        
        # Trova date backup (veloce)
        all_dates = self.get_backup_dates_fast(verbose)
        self.stats['total_records'] = len(all_dates)
        self.stats['_all_dates'] = all_dates  # Salva per export
        
        if all_dates:
            self.stats['newest_backup'] = all_dates[0]['datetime']
            self.stats['oldest_backup'] = all_dates[-1]['datetime']
        
        return self.stats
    
    def load_all(self, verbose: bool = True, calc_sizes: bool = False, max_records: int = 20):
        """Carica tutti i dati del backup set."""
        self.calculate_sizes = calc_sizes
        self.max_records_per_folder = max_records
        
        if verbose:
            print(f"  📂 {self.name}")
        
        self.load_config()
        self.load_plan()
        self.load_folders(verbose)
        self.calculate_stats(verbose)
        
        if verbose:
            status = "✅" if self.stats['folders_count'] > 0 else "⚠️"
            encrypted = "🔒" if self.stats.get('is_encrypted') else "🔓"
            print(f"     {status} {encrypted} {self.stats['folders_count']} cartelle, {self.stats['total_records']} records")
            if self.stats.get('newest_backup'):
                print(f"     📅 Ultimo backup: {self.stats['newest_backup'].strftime('%Y-%m-%d %H:%M')}")
            print()


class ArqAnalyzer:
    """Analizzatore principale per tutti i backup Arq."""
    
    def __init__(self, base_path: str):
        self.base_path = Path(base_path)
        self.backup_sets: Dict[str, ArqBackupSet] = {}
        self.comparison: List[Dict] = []
        
    def load_all_backups(self, verbose: bool = True, calc_sizes: bool = False, 
                         max_records: int = 20) -> Dict[str, ArqBackupSet]:
        """Carica tutti i backup set configurati."""
        if verbose:
            print("\n" + "=" * 70)
            print("  ANALISI BACKUP ARQ v2.1")
            print("=" * 70 + "\n")
        
        for uuid, name in MAC_MAPPING.items():
            backup_set = ArqBackupSet(self.base_path, uuid, name)
            
            if backup_set.exists():
                backup_set.load_all(verbose, calc_sizes, max_records)
            else:
                if verbose:
                    print(f"  ⚠️  {name}: backup non trovato")
                    print()
            
            self.backup_sets[name] = backup_set
        
        return self.backup_sets
    
    def compare_folders(self) -> List[Dict]:
        """Confronta le cartelle tra i diversi Mac."""
        all_paths = {}
        
        for mac_name, backup_set in self.backup_sets.items():
            for folder in backup_set.folders:
                path = folder.get('localPath', '')
                if path not in all_paths:
                    all_paths[path] = {
                        'path': path,
                        'name': folder.get('name', 'Unknown'),
                        'macs': [],
                        'details': {}
                    }
                all_paths[path]['macs'].append(mac_name)
                all_paths[path]['details'][mac_name] = {
                    'uuid': folder.get('uuid') or folder.get('_uuid'),
                    'storage_class': folder.get('storageClass', 'STANDARD')
                }
        
        self.comparison = []
        for path, data in sorted(all_paths.items()):
            count = len(data['macs'])
            if count == 3:
                category = '🟢 Comune a tutti'
            elif count == 2:
                category = '🟡 Comune a 2'
            else:
                category = '🔴 Unico'
            
            data['category'] = category
            data['count'] = count
            data['user'] = extract_user_from_path(path)
            self.comparison.append(data)
        
        self.comparison.sort(key=lambda x: (-x['count'], x['path']))
        
        return self.comparison
    
    def print_summary(self):
        """Stampa riepilogo a console."""
        print("\n" + "─" * 70)
        print("  RIEPILOGO")
        print("─" * 70 + "\n")
        
        for mac_name, backup_set in self.backup_sets.items():
            stats = backup_set.stats
            if not backup_set.exists():
                print(f"❌ {mac_name}: non trovato\n")
                continue
                
            status = "✅" if stats.get('folders_count', 0) > 0 else "⚠️"
            encrypted = "🔒 Criptato" if stats.get('is_encrypted') else "🔓 Non criptato"
            
            print(f"{status} {mac_name}")
            print(f"   Computer: {stats.get('computer_name', '-')}")
            print(f"   Stato: {encrypted}")
            print(f"   Cartelle: {stats.get('folders_count', 0)}")
            print(f"   Records: {stats.get('total_records', 0)}")
            print(f"   Dimensione: {stats.get('total_size_human', 'N/D')}")
            
            if stats.get('newest_backup'):
                print(f"   Ultimo backup: {stats['newest_backup'].strftime('%Y-%m-%d %H:%M')}")
            if stats.get('oldest_backup'):
                print(f"   Primo backup: {stats['oldest_backup'].strftime('%Y-%m-%d %H:%M')}")
            print()
        
        # Statistiche confronto
        common_all = sum(1 for c in self.comparison if c['count'] == 3)
        common_two = sum(1 for c in self.comparison if c['count'] == 2)
        unique = sum(1 for c in self.comparison if c['count'] == 1)
        
        print("─" * 70)
        print("  CONFRONTO CARTELLE")
        print("─" * 70 + "\n")
        print(f"  🟢 Comuni a tutti e 3 Mac: {common_all}")
        print(f"  🟡 Comuni a 2 Mac: {common_two}")
        print(f"  🔴 Uniche (1 solo Mac): {unique}")
        print(f"\n  Totale cartelle: {len(self.comparison)}")
    
    def print_comparison_details(self, limit: int = 30):
        """Stampa dettagli confronto."""
        print("\n" + "─" * 70)
        print("  DETTAGLIO CARTELLE")
        print("─" * 70 + "\n")
        
        for idx, comp in enumerate(self.comparison[:limit], 1):
            print(f"{idx:3}. {comp['category']} | {comp['name']}")
            print(f"     Path: {comp['path']}")
            print(f"     Su: {', '.join(comp['macs'])}")
            print()
        
        if len(self.comparison) > limit:
            print(f"  ... e altre {len(self.comparison) - limit} cartelle (vedi Excel)")
    
    def export_excel(self, output_path: str) -> bool:
        """Esporta report Excel."""
        if not HAS_OPENPYXL:
            print("⚠️  Export Excel non disponibile (installa openpyxl)")
            return False
        
        print(f"\n📊 Creazione report Excel...")
        
        wb = openpyxl.Workbook()
        
        # Stili
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # ===== FOGLIO 1: Riepilogo =====
        ws = wb.active
        ws.title = "Riepilogo"
        
        ws['A1'] = 'ANALISI BACKUP ARQ'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:G1')
        
        ws['A2'] = f'Generato: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        
        row = 4
        headers = ['Mac', 'Computer', 'Cartelle', 'Records', 'Ultimo Backup', 'Criptato']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        row += 1
        for mac_name, backup_set in self.backup_sets.items():
            stats = backup_set.stats
            ws.cell(row=row, column=1, value=mac_name).border = thin_border
            ws.cell(row=row, column=2, value=stats.get('computer_name', '-')).border = thin_border
            ws.cell(row=row, column=3, value=stats.get('folders_count', 0)).border = thin_border
            ws.cell(row=row, column=4, value=stats.get('total_records', 0)).border = thin_border
            
            newest = stats.get('newest_backup')
            ws.cell(row=row, column=5, value=newest.strftime('%Y-%m-%d %H:%M') if newest else '-').border = thin_border
            ws.cell(row=row, column=6, value='Sì' if stats.get('is_encrypted') else 'No').border = thin_border
            
            fill = green_fill if stats.get('folders_count', 0) > 0 else red_fill
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = fill
            row += 1
        
        # Stats
        row += 2
        common_all = sum(1 for c in self.comparison if c['count'] == 3)
        common_two = sum(1 for c in self.comparison if c['count'] == 2)
        unique = sum(1 for c in self.comparison if c['count'] == 1)
        
        ws.cell(row=row, column=1, value='Comuni a tutti:')
        ws.cell(row=row, column=2, value=common_all).fill = green_fill
        row += 1
        ws.cell(row=row, column=1, value='Comuni a 2:')
        ws.cell(row=row, column=2, value=common_two).fill = yellow_fill
        row += 1
        ws.cell(row=row, column=1, value='Uniche:')
        ws.cell(row=row, column=2, value=unique).fill = red_fill
        
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # ===== FOGLIO 2: Confronto =====
        ws2 = wb.create_sheet("Confronto")
        
        headers = ['Categoria', 'Percorso', 'Nome', 'Utente', 'Presente su']
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, item in enumerate(self.comparison, 2):
            ws2.cell(row=row_idx, column=1, value=item['category'])
            ws2.cell(row=row_idx, column=2, value=item['path'])
            ws2.cell(row=row_idx, column=3, value=item['name'])
            ws2.cell(row=row_idx, column=4, value=item['user'])
            ws2.cell(row=row_idx, column=5, value=', '.join(item['macs']))
            
            fill = green_fill if item['count'] == 3 else (yellow_fill if item['count'] == 2 else red_fill)
            for col in range(1, 6):
                ws2.cell(row=row_idx, column=col).fill = fill
        
        ws2.column_dimensions['A'].width = 18
        ws2.column_dimensions['B'].width = 50
        ws2.column_dimensions['C'].width = 25
        ws2.column_dimensions['D'].width = 15
        ws2.column_dimensions['E'].width = 40
        
        # ===== FOGLI per Mac =====
        for mac_name, backup_set in self.backup_sets.items():
            sheet_name = mac_name[:31]
            ws_mac = wb.create_sheet(sheet_name)
            
            ws_mac['A1'] = f'{mac_name}'
            ws_mac['A1'].font = Font(bold=True, size=14)
            
            headers = ['#', 'Nome', 'Percorso', 'Utente']
            for col, header in enumerate(headers, 1):
                cell = ws_mac.cell(row=3, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            folders_sorted = sorted(backup_set.folders, key=lambda x: x.get('localPath', ''))
            for idx, folder in enumerate(folders_sorted, 1):
                row = idx + 3
                ws_mac.cell(row=row, column=1, value=idx)
                ws_mac.cell(row=row, column=2, value=folder.get('name', ''))
                ws_mac.cell(row=row, column=3, value=folder.get('localPath', ''))
                ws_mac.cell(row=row, column=4, value=extract_user_from_path(folder.get('localPath', '')))
            
            ws_mac.column_dimensions['A'].width = 5
            ws_mac.column_dimensions['B'].width = 30
            ws_mac.column_dimensions['C'].width = 50
            ws_mac.column_dimensions['D'].width = 15
        
        # ===== FOGLIO Timeline =====
        ws_time = wb.create_sheet("Timeline")
        
        headers = ['Mac', 'Cartella', 'Data', 'Ora']
        for col, header in enumerate(headers, 1):
            cell = ws_time.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        row = 2
        for mac_name, backup_set in self.backup_sets.items():
            all_dates = backup_set.stats.get('_all_dates', [])[:50]
            for d in all_dates:
                ws_time.cell(row=row, column=1, value=mac_name)
                ws_time.cell(row=row, column=2, value=d['folder_name'])
                ws_time.cell(row=row, column=3, value=d['datetime'].strftime('%Y-%m-%d'))
                ws_time.cell(row=row, column=4, value=d['datetime'].strftime('%H:%M'))
                row += 1
        
        ws_time.column_dimensions['A'].width = 15
        ws_time.column_dimensions['B'].width = 30
        ws_time.column_dimensions['C'].width = 12
        ws_time.column_dimensions['D'].width = 10
        
        # Salva
        wb.save(output_path)
        print(f"✅ Report salvato: {output_path}")
        
        return True


def main():
    parser = argparse.ArgumentParser(
        description='Arq Backup Analyzer v2.1 - Analizza backup Arq 7',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    parser.add_argument('path', nargs='?', default=DEFAULT_ARQ_PATH,
                        help='Percorso Arq Backup Data')
    parser.add_argument('-o', '--output', help='File Excel output')
    parser.add_argument('-d', '--details', action='store_true',
                        help='Mostra dettaglio cartelle')
    parser.add_argument('-s', '--calc-size', action='store_true',
                        help='Calcola dimensioni (lento!)')
    parser.add_argument('-m', '--max-records', type=int, default=20,
                        help='Max records per cartella (default: 20)')
    parser.add_argument('-q', '--quiet', action='store_true',
                        help='Output minimo')
    
    args = parser.parse_args()
    log.info(f"Avviato v{VERSION} — path: {args.path}")

    if not Path(args.path).exists():
        print(f"❌ Percorso non trovato: {args.path}")
        sys.exit(1)
    
    # Analizza
    analyzer = ArqAnalyzer(args.path)
    analyzer.load_all_backups(
        verbose=not args.quiet, 
        calc_sizes=args.calc_size,
        max_records=args.max_records
    )
    analyzer.compare_folders()
    
    if not args.quiet:
        analyzer.print_summary()
        if args.details:
            analyzer.print_comparison_details()
    
    # Export
    output_path = args.output or str(app_output_dir(APP_NAME) / f"Arq_Analysis_{timestamp}.xlsx")
    
    if HAS_OPENPYXL:
        analyzer.export_excel(output_path)
    log.info(f"Analisi completata — output: {output_path}")

    print("\n" + "=" * 70)
    print("✅ ANALISI COMPLETATA")
    print("=" * 70)
    print(f"\n📄 Log: {log.log_path}")
    if HAS_OPENPYXL:
        import subprocess
        subprocess.run(["open", output_path])


if __name__ == "__main__":
    main()